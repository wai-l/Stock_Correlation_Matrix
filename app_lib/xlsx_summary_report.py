# import pandas as pd
# import io
# from openpyxl.styles import Font, Alignment, Border

# def xlsx_summary_report(price_history, corr_matrix):
#     # the buffer is to let the excel file be the returning value of the function
#     buffer = io.BytesIO()
#     with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
#         price_history.to_excel(writer, sheet_name='price_history', index=False, header=True)
#         corr_matrix.to_excel(writer, sheet_name='correlation_matrix', index=True, header=True)
#         worksheet1 = writer.sheets['price_history']
#         worksheet2 = writer.sheets['correlation_matrix']

#     # Apply formatting to first row and first column
#         for worksheet in [worksheet1, worksheet2]:
#             # first row
#             for cell in worksheet['1:1']:
#                 cell.font = Font(bold=False)
#                 cell.alignment = Alignment(horizontal='left', vertical='center')
#                 cell.border = None
#             # first column
#             for row in worksheet.iter_rows(min_row=2, min_col=1):
#                 for cell in row:
#                     cell.alignment = Alignment(horizontal='left', vertical='center')
#                     cell.font = Font(bold=False)
#                     cell.border = None

#     return buffer

import io
from dataclasses import dataclass
from typing import Callable, Optional, Sequence, Any, Dict

import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter



# ----------------------------
# Formatting helpers
# ----------------------------
def default_sheet_formatter(ws):
    """
    Keep formatting conservative and consistent.
    - Header row: left aligned, not bold
    - Data cells: left aligned, not bold
    """
    header_font = Font(bold=False)
    left_align = Alignment(horizontal="left", vertical="center")

    # Header row
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = left_align
        cell.border = None

    # Data area
    # (avoid styling an entire column if the sheet is empty)
    if ws.max_row >= 2 and ws.max_column >= 1:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = header_font
                cell.alignment = left_align
                cell.border = None

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0

        for cell in ws[col_letter]:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)

        if max_length > 0:
            # Padding factor (Excel ≈ 1.1–1.2)
            ws.column_dimensions[col_letter].width = max_length + 2


@dataclass(frozen=True)
class SheetSpec:
    name: str
    df: pd.DataFrame
    index: bool = False
    header: bool = True
    formatter: Optional[Callable[[Any], None]] = default_sheet_formatter
    to_excel_kwargs: Optional[Dict[str, Any]] = None  # for future extensions


# ----------------------------
# Main export function
# ----------------------------
def xlsx_summary_report(
    sheets: Sequence[SheetSpec]
    ) -> io.BytesIO:
    """
    Build an in-memory Excel workbook with multiple sheets.
    Easy to expand: just add another SheetSpec to `sheets`.
    Returns a BytesIO buffer suitable for Streamlit download_button.
    """
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # 1) Write sheets
        for spec in sheets:
            kwargs = spec.to_excel_kwargs or {}
            spec.df.to_excel(
                writer,
                sheet_name=spec.name,
                index=spec.index,
                header=spec.header,
                **kwargs
            )

        # 2) Apply per-sheet formatting
        for spec in sheets:
            ws = writer.sheets[spec.name]
            if spec.formatter:
                spec.formatter(ws)

    buffer.seek(0)
    return buffer


# ----------------------------
# Convenience wrapper for your current use-case
# ----------------------------
def build_portfolio_export(
        para: pd.DataFrame, 
        portfolio_allocation: pd.DataFrame, 
        portfo_summary: pd.DataFrame, 
        asset_metric: pd.DataFrame, 
        asset_contrib: pd.DataFrame, 
        price_history: pd.DataFrame, 
        price_history_indexed: pd.DataFrame, 
        corr_matrix: pd.DataFrame
        ) -> io.BytesIO:
    sheets = [
        SheetSpec(name='parameters', df=para, index=False, header=True), 
        SheetSpec(name='portfolio_allocation', df = portfolio_allocation, index=False, header=True), 
        SheetSpec(name='portfo_summary', df=portfo_summary, index=False, header=True), 
        SheetSpec(name='asset_metric', df=asset_metric, index=True, header=True), 
        SheetSpec(name='asset_contrib', df=asset_contrib, index=False, header=True), 
        SheetSpec(name="price_history", df=price_history, index=False, header=True),
        SheetSpec(name="price_history_indexed", df=price_history_indexed, index=False, header=True), 
        SheetSpec(name="correlation_matrix", df=corr_matrix, index=True, header=True)
    ]
    return xlsx_summary_report(sheets)
